"""
batch_eval.py — Batch accuracy evaluation for the Monixor 2.0 pipeline.

Usage:
    1. Make sure pipeline_api.py is running (python pipeline_api.py)
    2. python batch_eval.py

Scans DATASET_DIR for images, matches each to the ground-truth CSV by
filename stem, runs the pipeline, and writes a colour-coded Excel report
(batch_results.xlsx) plus a plain CSV backup (batch_results.csv).
"""

import csv
import io
import os
import time
import urllib.request

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuration ──────────────────────────────────────────────────────────────
FLASK_URL   = "http://localhost:5000/run_pipeline"

DATASET_DIR = r"C:\Users\Tela\Documents\Thesis\Dataset"
GT_CSV_NAME = "Monixor2.0_Ground Truth - Sheet1.csv"

# Path to main-branch batch results CSV for side-by-side comparison.
# Set to None to skip the comparison section.
MAIN_BRANCH_CSV = r"C:\Users\Tela\Downloads\Monixor2.0_Batch_Results.csv"

# Main-branch CSV key names that map to dev-branch vital keys
MAIN_KEY_MAP = {
    "HR": "HR", "SpO2": "SpO2", "PR": "PR",
    "RR": "Resp", "NIBP": "NIBP", "MAP": "MAP", "Temp": "Temp",
}

GT_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "1FFgE1AiqdIHChd5Tf83lhl43m8DiAsEoOs8pykgtXXA"
    "/export?format=csv"
)

OUTPUT_XLSX  = "Left_batch_results.xlsx"
OUTPUT_CSV   = "Left_batch_results.csv"
IMAGE_EXTS   = {".jpg", ".jpeg", ".png"}
LIMIT        = None   # e.g. 10 to run only first N; None = all
ANGLE_FILTER = "Left"  # e.g. "Left" or ["Left","Right"] to test specific angles only; None = all
TOLERANCE    = 1

VITALS = [
    ("HR",   "HR"),
    ("SpO2", "Sp02"),
    ("PR",   "PR"),
    ("RR",   "Resp"),
    ("NIBP", "NIBP"),
    ("MAP",  "MAP"),
    ("Temp", "Temp"),
]

# ── Colours ────────────────────────────────────────────────────────────────────
CLR_HEADER      = "1F4E79"   # dark blue
CLR_SUBHEADER   = "2E75B6"   # medium blue
CLR_CORRECT     = "C6EFCE"   # light green
CLR_WRONG       = "FFC7CE"   # light red/pink
CLR_GT          = "DDEBF7"   # light blue (ground truth cells)
CLR_SCORE_FULL  = "70AD47"   # green  (all correct)
CLR_SCORE_PART  = "FFD966"   # yellow (some wrong)
CLR_SCORE_ZERO  = "FF0000"   # red    (all wrong)
CLR_SUMMARY_BG  = "F2F2F2"   # light grey
CLR_WHITE       = "FFFFFF"
# ──────────────────────────────────────────────────────────────────────────────


def load_ground_truth():
    local_path = os.path.join(DATASET_DIR, GT_CSV_NAME)
    if os.path.exists(local_path):
        print(f"Loading ground truth from local file: {local_path}")
        with open(local_path, encoding="utf-8") as f:
            raw = f.read()
    else:
        print("Local GT CSV not found — loading from Google Sheets…")
        raw = urllib.request.urlopen(GT_URL).read().decode("utf-8")

    rows = list(csv.DictReader(io.StringIO(raw)))
    rows = [{k.strip(): v.strip() for k, v in row.items()} for row in rows]
    print(f"  {len(rows)} ground-truth rows loaded.\n")

    gt_lookup = {}
    for row in rows:
        stem = row.get("Image", "").strip().lower()
        if stem:
            gt_lookup[stem] = row
    return gt_lookup


def scan_images(folder):
    images = []
    for fname in sorted(os.listdir(folder)):
        ext = os.path.splitext(fname)[1].lower()
        if ext in IMAGE_EXTS:
            images.append(os.path.join(folder, fname))
    return images


def is_correct(pred_str, gt_str, vital_key):
    pred_str = str(pred_str).strip()
    gt_str   = str(gt_str).strip()
    if not pred_str or not gt_str:
        return False
    if vital_key == "NIBP":
        return pred_str == gt_str
    try:
        return abs(float(pred_str) - float(gt_str)) <= TOLERANCE
    except ValueError:
        return pred_str == gt_str


def _fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def _font(bold=False, color=CLR_WHITE, size=11):
    return Font(bold=bold, color=color, size=size)

def _border():
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def build_excel(results, total_images):
    wb = Workbook()
    ws = wb.active
    ws.title = "Batch Results"

    n_vitals = len(VITALS)

    # ── Row 1: main header ─────────────────────────────────────────────────
    # Columns layout:
    #   A: #  B: Image  C: Session  D: Angle  E: Time(s)
    #   Then for each vital: GT | Pred   (2 cols each)
    #   Last col: Score
    meta_cols  = 5
    vital_cols = n_vitals * 2   # GT + Pred per vital
    score_col  = meta_cols + vital_cols + 1
    total_cols = score_col

    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1,   end_column=meta_cols)
    hdr_cell = ws.cell(row=1, column=1, value="Monixor 2.0 — Batch Evaluation Results")
    hdr_cell.font      = Font(bold=True, color=CLR_WHITE, size=13)
    hdr_cell.fill      = _fill(CLR_HEADER)
    hdr_cell.alignment = _center()

    ws.merge_cells(start_row=1, start_column=meta_cols+1,
                   end_row=1,   end_column=meta_cols+vital_cols)
    vhdr = ws.cell(row=1, column=meta_cols+1,
                   value=f"Vital Signs  (Tolerance ±{TOLERANCE} for numeric, exact for NIBP)")
    vhdr.font      = Font(bold=True, color=CLR_WHITE, size=12)
    vhdr.fill      = _fill(CLR_HEADER)
    vhdr.alignment = _center()

    score_hdr = ws.cell(row=1, column=score_col, value="Score")
    score_hdr.font      = Font(bold=True, color=CLR_WHITE, size=12)
    score_hdr.fill      = _fill(CLR_HEADER)
    score_hdr.alignment = _center()

    # ── Row 2: vital-name sub-headers ──────────────────────────────────────
    for c in range(1, meta_cols+1):
        ws.cell(row=2, column=c).fill = _fill(CLR_SUBHEADER)

    for i, (api_key, _) in enumerate(VITALS):
        col_gt   = meta_cols + i*2 + 1
        col_pred = meta_cols + i*2 + 2
        ws.merge_cells(start_row=2, start_column=col_gt,
                       end_row=2,   end_column=col_pred)
        c = ws.cell(row=2, column=col_gt, value=api_key)
        c.font      = Font(bold=True, color=CLR_WHITE, size=11)
        c.fill      = _fill(CLR_SUBHEADER)
        c.alignment = _center()

    ws.cell(row=2, column=score_col).fill = _fill(CLR_SUBHEADER)

    # ── Row 3: column labels ───────────────────────────────────────────────
    labels = ["#", "Image", "Session", "Angle", "Time (s)"]
    for col, lbl in enumerate(labels, 1):
        c = ws.cell(row=3, column=col, value=lbl)
        c.font      = Font(bold=True, color="1F4E79", size=10)
        c.fill      = _fill("D6E4F7")
        c.alignment = _center()
        c.border    = _border()

    for i, (api_key, _) in enumerate(VITALS):
        col_gt   = meta_cols + i*2 + 1
        col_pred = meta_cols + i*2 + 2
        for col, lbl in [(col_gt, "GT"), (col_pred, "Pred")]:
            c = ws.cell(row=3, column=col, value=lbl)
            c.font      = Font(bold=True, color="1F4E79", size=10)
            c.fill      = _fill("D6E4F7")
            c.alignment = _center()
            c.border    = _border()

    sc = ws.cell(row=3, column=score_col, value="Score")
    sc.font      = Font(bold=True, color="1F4E79", size=10)
    sc.fill      = _fill("D6E4F7")
    sc.alignment = _center()
    sc.border    = _border()

    # ── Data rows ─────────────────────────────────────────────────────────
    for idx, rec in enumerate(results, 1):
        data_row = idx + 3
        correct_count = sum(1 for api_key, _ in VITALS if rec[f"{api_key}_ok"] == "1")
        score_str = f"{correct_count}/{n_vitals}"

        # Meta columns
        meta_vals = [idx, rec["Image"], rec["Session"], rec["Angle"], rec["Time_s"]]
        for col, val in enumerate(meta_vals, 1):
            c = ws.cell(row=data_row, column=col, value=val)
            c.alignment = _center()
            c.border    = _border()
            row_bg = "FFFFFF" if idx % 2 == 0 else "F7FBFF"
            c.fill = _fill(row_bg)

        # Vital columns
        for i, (api_key, _) in enumerate(VITALS):
            col_gt   = meta_cols + i*2 + 1
            col_pred = meta_cols + i*2 + 2
            ok       = rec[f"{api_key}_ok"] == "1"

            gt_cell = ws.cell(row=data_row, column=col_gt,
                              value=rec[f"{api_key}_gt"])
            gt_cell.fill      = _fill(CLR_GT)
            gt_cell.alignment = _center()
            gt_cell.border    = _border()

            pred_cell = ws.cell(row=data_row, column=col_pred,
                                value=rec[f"{api_key}_pred"])
            pred_cell.fill      = _fill(CLR_CORRECT if ok else CLR_WRONG)
            pred_cell.font      = Font(bold=not ok, color="000000", size=11)
            pred_cell.alignment = _center()
            pred_cell.border    = _border()

        # Score column
        if correct_count == n_vitals:
            score_fill = CLR_SCORE_FULL
        elif correct_count == 0:
            score_fill = CLR_SCORE_ZERO
        else:
            score_fill = CLR_SCORE_PART

        sc = ws.cell(row=data_row, column=score_col, value=score_str)
        sc.fill      = _fill(score_fill)
        sc.font      = Font(bold=True, color="000000", size=11)
        sc.alignment = _center()
        sc.border    = _border()

    # ── Summary rows ──────────────────────────────────────────────────────
    summary_row = len(results) + 4 + 1  # blank row gap

    ws.cell(row=summary_row, column=1, value="ACCURACY SUMMARY").font = Font(
        bold=True, size=11, color="1F4E79")

    # Per-vital accuracy
    acc_row = summary_row + 1
    label_cell = ws.cell(row=acc_row, column=1, value="Accuracy %")
    label_cell.font      = Font(bold=True, size=10)
    label_cell.fill      = _fill(CLR_SUMMARY_BG)
    label_cell.alignment = _center()
    label_cell.border    = _border()
    # blank meta cols 2-5
    for col in range(2, meta_cols+1):
        c = ws.cell(row=acc_row, column=col)
        c.fill   = _fill(CLR_SUMMARY_BG)
        c.border = _border()

    n = len(results)
    for i, (api_key, _) in enumerate(VITALS):
        col_gt   = meta_cols + i*2 + 1
        col_pred = meta_cols + i*2 + 2
        correct  = sum(1 for r in results if r[f"{api_key}_ok"] == "1")
        pct      = correct / n * 100 if n else 0

        gt_c = ws.cell(row=acc_row, column=col_gt,
                       value=f"{correct}/{n}")
        gt_c.fill      = _fill(CLR_SUMMARY_BG)
        gt_c.alignment = _center()
        gt_c.border    = _border()

        pred_c = ws.cell(row=acc_row, column=col_pred,
                         value=f"{pct:.1f}%")
        pred_c.font      = Font(bold=True, color="000000")
        pred_c.fill      = _fill(CLR_SCORE_FULL if pct == 100 else
                                 (CLR_SCORE_PART if pct >= 70 else CLR_SCORE_ZERO))
        pred_c.alignment = _center()
        pred_c.border    = _border()

    # All-correct row
    all_row   = acc_row + 1
    all_ok    = sum(1 for r in results
                    if all(r[f"{api_key}_ok"] == "1" for api_key, _ in VITALS))
    all_pct   = all_ok / n * 100 if n else 0

    ws.merge_cells(start_row=all_row, start_column=1,
                   end_row=all_row,   end_column=meta_cols)
    all_lbl = ws.cell(row=all_row, column=1, value="All Correct")
    all_lbl.font      = Font(bold=True, size=10, color=CLR_WHITE)
    all_lbl.fill      = _fill(CLR_HEADER)
    all_lbl.alignment = _center()
    all_lbl.border    = _border()

    ws.merge_cells(start_row=all_row, start_column=meta_cols+1,
                   end_row=all_row,   end_column=score_col)
    all_val = ws.cell(row=all_row, column=meta_cols+1,
                      value=f"{all_ok} / {n}  ({all_pct:.1f}%)")
    all_val.font      = Font(bold=True, size=12, color=CLR_WHITE)
    all_val.fill      = _fill(CLR_SCORE_FULL if all_pct >= 90 else
                               (CLR_SCORE_PART if all_pct >= 70 else CLR_SCORE_ZERO))
    all_val.alignment = _center()
    all_val.border    = _border()

    # ── Legend ─────────────────────────────────────────────────────────────
    leg_row = all_row + 2
    ws.cell(row=leg_row, column=1, value="Legend:").font = Font(bold=True, size=9)
    legend = [
        (CLR_CORRECT,    "Correct prediction"),
        (CLR_WRONG,      "Wrong prediction"),
        (CLR_GT,         "Ground truth value"),
        (CLR_SCORE_FULL, "All vitals correct"),
        (CLR_SCORE_PART, "Some vitals wrong"),
    ]
    for j, (color, label) in enumerate(legend):
        col = j * 2 + 1
        box = ws.cell(row=leg_row+1, column=col, value="  ")
        box.fill   = _fill(color)
        box.border = _border()
        txt = ws.cell(row=leg_row+1, column=col+1, value=label)
        txt.font = Font(size=9)

    # ── Column widths ──────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 5   # #
    ws.column_dimensions["B"].width = 14  # Image
    ws.column_dimensions["C"].width = 9   # Session
    ws.column_dimensions["D"].width = 9   # Angle
    ws.column_dimensions["E"].width = 8   # Time

    for i in range(vital_cols):
        col_letter = get_column_letter(meta_cols + i + 1)
        ws.column_dimensions[col_letter].width = 9

    ws.column_dimensions[get_column_letter(score_col)].width = 8

    # Freeze panes (keep headers visible while scrolling)
    ws.freeze_panes = ws.cell(row=4, column=meta_cols+1)

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 16

    wb.save(OUTPUT_XLSX)
    print(f"Excel report saved to: {OUTPUT_XLSX}")


def run_batch():
    gt_lookup = load_ground_truth()
    images    = scan_images(DATASET_DIR)

    if not images:
        print(f"No images found in {DATASET_DIR}")
        return

    if LIMIT:
        images = images[:LIMIT]
        print(f"Running on first {LIMIT} images (LIMIT={LIMIT}).\n")
    else:
        print(f"Found {len(images)} image(s) in folder.\n")

    results   = []
    skipped   = 0
    no_gt     = 0
    processed = 0

    csv_fieldnames = (
        ["Session", "Angle", "Image", "Status", "Time_s"]
        + [f"{api_key}_gt"   for api_key, _ in VITALS]
        + [f"{api_key}_pred" for api_key, _ in VITALS]
        + [f"{api_key}_ok"   for api_key, _ in VITALS]
    )

    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as out_f:
        writer = csv.DictWriter(out_f, fieldnames=csv_fieldnames)
        writer.writeheader()

        for idx, img_path in enumerate(images, 1):
            fname  = os.path.basename(img_path)
            stem   = os.path.splitext(fname)[0]
            gt_row = gt_lookup.get(stem.lower())

            print(f"[{idx:>3}/{len(images)}] {fname} … ", end="", flush=True)

            if gt_row is None:
                print("SKIP — no ground truth in CSV")
                no_gt += 1
                continue

            angle   = gt_row.get("Angle", "").strip()
            session = gt_row.get("Session", "").strip()

            if ANGLE_FILTER:
                allowed = [ANGLE_FILTER] if isinstance(ANGLE_FILTER, str) else ANGLE_FILTER
                if angle not in allowed:
                    print(f"SKIP — angle '{angle}' not in filter")
                    continue

            t0 = time.time()
            try:
                with open(img_path, "rb") as img_f:
                    resp = requests.post(
                        FLASK_URL,
                        files={"monitor_photo": img_f},
                        timeout=180,
                    )
                elapsed = round(time.time() - t0, 1)

                if resp.status_code != 200:
                    print(f"ERR HTTP {resp.status_code} ({elapsed}s)")
                    skipped += 1
                    continue

                predicted = resp.json()

            except requests.exceptions.Timeout:
                print("ERR — timeout")
                skipped += 1
                continue
            except Exception as exc:
                print(f"ERR — {exc}")
                skipped += 1
                continue

            record = {
                "Session": session,
                "Angle":   angle,
                "Image":   stem,
                "Status":  "OK",
                "Time_s":  elapsed,
            }

            correct_count = 0
            for api_key, gt_key in VITALS:
                gt_val   = gt_row.get(gt_key, "").strip()
                pred_val = str(predicted.get(api_key, "") or "").strip()
                ok       = is_correct(pred_val, gt_val, api_key)
                if ok:
                    correct_count += 1
                record[f"{api_key}_gt"]   = gt_val
                record[f"{api_key}_pred"] = pred_val
                record[f"{api_key}_ok"]   = "1" if ok else "0"

            results.append(record)
            writer.writerow(record)
            out_f.flush()
            processed += 1

            vital_str = "  ".join(
                f"{api_key}={'✓' if record[f'{api_key}_ok']=='1' else '✗'}"
                f"({record[f'{api_key}_pred']}|gt:{record[f'{api_key}_gt']})"
                for api_key, _ in VITALS
            )
            print(f"OK ({elapsed}s)  {correct_count}/{len(VITALS)}  —  {vital_str}")

    print(f"\n{'='*70}")
    print(f"Processed: {processed}  |  No GT: {no_gt}  |  Errors/Skipped: {skipped}")
    print(f"{'='*70}\n")

    if not results:
        print("No results to summarise.")
        return

    # ── Console accuracy summary ───────────────────────────────────────────
    n = len(results)

    # Load main-branch results for same images (if CSV configured)
    main_lookup = {}
    if MAIN_BRANCH_CSV and os.path.exists(MAIN_BRANCH_CSV):
        with open(MAIN_BRANCH_CSV, encoding="utf-8") as mf:
            for row in csv.DictReader(mf):
                stem = os.path.splitext(row.get("image", "").strip())[0].lower()
                if stem:
                    main_lookup[stem] = row

    # Build per-vital main stats for the images we just processed
    main_key = {"HR":"HR","SpO2":"SpO2","PR":"PR","RR":"Resp",
                "NIBP":"NIBP","MAP":"MAP","Temp":"Temp"}
    main_rows = [main_lookup[r["Image"].lower()]
                 for r in results if r["Image"].lower() in main_lookup]
    has_main  = len(main_rows) > 0

    print("\nACCURACY SUMMARY")
    print("─" * 60)
    if has_main:
        print(f"{'Vital':<14}  {'Dev':>12}  {'Main (ref)':>12}  {'Δ':>5}")
        print(f"{'':14}  {'correct / %':>12}  {'correct / %':>12}  {'':>5}")
    else:
        print(f"{'Vital':<14}  {'Correct':>7}  {'%':>7}")
    print("─" * 60)

    dev_total = 0
    main_total = 0
    for api_key, _ in VITALS:
        dev_ok  = sum(1 for r in results if r[f"{api_key}_ok"] == "1")
        dev_pct = dev_ok / n * 100
        dev_total += dev_ok
        label = api_key if api_key == "NIBP" else f"{api_key} (±{TOLERANCE})"
        if has_main:
            mk = main_key[api_key]
            main_ok  = sum(1 for mr in main_rows
                           if str(mr.get(f"{mk}_ok","")).strip().upper() == "TRUE")
            main_pct = main_ok / len(main_rows) * 100
            main_total += main_ok
            delta = dev_pct - main_pct
            sign  = "+" if delta >= 0 else ""
            print(f"{label:<14}  {dev_ok:>3}/{n:<3} {dev_pct:>5.1f}%"
                  f"  {main_ok:>3}/{len(main_rows):<3} {main_pct:>5.1f}%"
                  f"  {sign}{delta:+.1f}%")
        else:
            print(f"{label:<14}  {dev_ok:>3}/{n:<3}  {dev_pct:>6.1f}%")

    dev_all  = sum(1 for r in results
                   if all(r[f"{api_key}_ok"] == "1" for api_key, _ in VITALS))
    dev_all_pct = dev_all / n * 100
    print("─" * 60)
    if has_main:
        m_all = sum(1 for mr in main_rows
                    if all(str(mr.get(f"{main_key[k]}_ok","")).strip().upper() == "TRUE"
                           for k, _ in VITALS))
        m_all_pct = m_all / len(main_rows) * 100
        delta_all = dev_all_pct - m_all_pct
        sign = "+" if delta_all >= 0 else ""
        print(f"{'All correct':<14}  {dev_all:>3}/{n:<3} {dev_all_pct:>5.1f}%"
              f"  {m_all:>3}/{len(main_rows):<3} {m_all_pct:>5.1f}%"
              f"  {sign}{delta_all:+.1f}%")
        dev_pv  = dev_total  / (n * len(VITALS)) * 100
        main_pv = main_total / (len(main_rows) * len(VITALS)) * 100
        delta_pv = dev_pv - main_pv
        sign = "+" if delta_pv >= 0 else ""
        print(f"{'Per-vital':<14}  {dev_total:>3}/{n*len(VITALS):<3} {dev_pv:>5.1f}%"
              f"  {main_total:>3}/{len(main_rows)*len(VITALS):<3} {main_pv:>5.1f}%"
              f"  {sign}{delta_pv:+.1f}%")
        print(f"\n  How 'Per-vital %' is computed:")
        print(f"    Dev : {dev_total} correct vitals ÷ ({n} images × {len(VITALS)} vitals)"
              f" = {dev_pv:.1f}%")
        print(f"    Main: {main_total} correct vitals ÷ ({len(main_rows)} images × {len(VITALS)} vitals)"
              f" = {main_pv:.1f}%")
    else:
        print(f"{'All correct':<14}  {dev_all:>3}/{n:<3}  {dev_all_pct:>6.1f}%")
        dev_pv = dev_total / (n * len(VITALS)) * 100
        print(f"{'Per-vital':<14}  {dev_total:>3}/{n*len(VITALS):<3}  {dev_pv:>6.1f}%")
    print()

    # ── Build Excel report ─────────────────────────────────────────────────
    build_excel(results, len(images))


if __name__ == "__main__":
    run_batch()
